from openpyxl import Workbook
import numpy as np

merge_file_name = "log-resultados"
events = ["10", "20", "30", "40", "50", "60", "70", "80", "90", "100", "110", "120", "130", "140", "150", "160", "170", "180", "190", "200"]


class RUN:
    def __init__(self, values):
        self.valores_ideales = values[0]
        self.costo_por_agente = values[1]
        self.completion_time_por_agente = values[2]
        self.tiempo_ultimo_agente_goal = values[3]
        self.tiempo_en_acabar = values[4]
        self.tiempo_promedio = values[5]
        self.agentes_en_goal = values[6]
        self.bad_good_total_rate_pred = values[7]
        self.push_out_count = values[8]


def strip_lines():
    lines = merge_file.readlines()
    for i in range(len(lines)):
        lines[i] = lines[i].strip()
    return lines


def read_instances(lines):
    RUNS = []
    temp = []
    arr_push_out = [0]
    for line in lines:
        name, values = line.split(' ')
        if name != "RUN":
            if name == "push_out_count":
                temp.append(eval(values) - arr_push_out[len(arr_push_out) - 1])
                arr_push_out.append(eval(values))
                # if run_name == 9:
                #     arr_push_out.clear()
                #     arr_push_out.append(0)
            else:
                temp.append(eval(values))
        if name == "push_out_count":
            RUNS.append(RUN(temp))
            temp.clear()
    return RUNS


def write_excel(RUNS, workbook):
    sheet = workbook.create_sheet(folder_name + "_agents")

    # tabla costo por agente #
    sheet.cell(row=1, column=1, value="costo_por_agente")
    sheet.cell(row=2, column=1, value="RUN")
    for i in range(int(folder_name)):
        sheet.cell(row=2, column=i + 2, value="agent " + str(i + 1))
    for i in range(len(RUNS)):
        sheet.cell(row=i + 3, column=1, value=str(i))
        RUN = RUNS[i]
        for j in range(len(RUN.costo_por_agente)):
            sheet.cell(row=i + 3, column=j + 2, value=RUN.costo_por_agente[j])
    initial_row = len(RUNS) + 4
    sheet.cell(row=initial_row - 1, column=1, value="promedio")
    sheet.cell(row=initial_row - 1, column=2, value=np.mean([np.mean(run.costo_por_agente) for run in RUNS]))
    sheet.cell(row=initial_row - 1, column=3, value="std")
    sheet.cell(row=initial_row - 1, column=4, value=np.std([np.mean(run.costo_por_agente) for run in RUNS]))

    # tabla completion time por agente#
    sheet.cell(row=initial_row, column=1, value="completion_time_por_agente")
    sheet.cell(row=initial_row + 1, column=1, value="RUN")
    for i in range(int(folder_name)):
        sheet.cell(row=initial_row + 1, column=i + 2, value="agent " + str(i + 1))
    for i in range(len(RUNS)):
        sheet.cell(row=initial_row + i + 2, column=1, value=str(i))
        RUN = RUNS[i]
        for j in range(len(RUN.completion_time_por_agente)):
            sheet.cell(row=initial_row + i + 2, column=j + 2, value=RUN.completion_time_por_agente[j])

    initial_row = (initial_row * 2) - 1
    sheet.cell(row=initial_row - 1, column=1, value="promedio")
    sheet.cell(row=initial_row - 1, column=2, value=np.mean([np.mean(run.completion_time_por_agente) for run in RUNS]))
    sheet.cell(row=initial_row - 1, column=3, value="std")
    sheet.cell(row=initial_row - 1, column=4, value=np.std([np.mean(run.completion_time_por_agente) for run in RUNS]))

    sheet.cell(row=initial_row, column=1, value="estadisticos")
    sheet.cell(row=initial_row + 1, column=1, value="RUN")
    sheet.cell(row=initial_row + 1, column=2, value="tiempo_ultimo_agente_goal")
    sheet.cell(row=initial_row + 1, column=3, value="tiempo_en_acabar")
    sheet.cell(row=initial_row + 1, column=4, value="tiempo_promedio")
    sheet.cell(row=initial_row + 1, column=5, value="agentes_en_goal")
    sheet.cell(row=initial_row + 1, column=6, value="bad_pred")
    sheet.cell(row=initial_row + 1, column=7, value="good_pred")
    sheet.cell(row=initial_row + 1, column=8, value="total_pred")
    sheet.cell(row=initial_row + 1, column=9, value="rate_pred")
    sheet.cell(row=initial_row + 1, column=10, value="push_out_count")
    for i in range(len(RUNS)):
        RUN = RUNS[i]
        sheet.cell(row=initial_row + i + 2, column=1, value=i)
        sheet.cell(row=initial_row + i + 2, column=2, value=RUN.tiempo_ultimo_agente_goal)
        sheet.cell(row=initial_row + i + 2, column=3, value=RUN.tiempo_en_acabar)
        sheet.cell(row=initial_row + i + 2, column=4, value=RUN.tiempo_promedio)
        sheet.cell(row=initial_row + i + 2, column=5, value=RUN.agentes_en_goal)
        sheet.cell(row=initial_row + i + 2, column=6, value=RUN.bad_good_total_rate_pred[0])
        sheet.cell(row=initial_row + i + 2, column=7, value=RUN.bad_good_total_rate_pred[1])
        sheet.cell(row=initial_row + i + 2, column=8, value=RUN.bad_good_total_rate_pred[2])
        sheet.cell(row=initial_row + i + 2, column=9, value=RUN.bad_good_total_rate_pred[3])
        sheet.cell(row=initial_row + i + 2, column=10, value=RUN.push_out_count)

    initial_row = initial_row + len(RUNS) + 3

    sheet.cell(row=initial_row, column=1, value="valores_ideales")
    for i in range(int(folder_name)):
        RUN = RUNS[0]
        sheet.cell(row=initial_row + 1, column=i + 1, value="agent " + str(i + 1))
        sheet.cell(row=initial_row + 2, column=i + 1, value=RUN.valores_ideales[i])
    sheet.cell(row=initial_row + 3, column=1, value="promedio_ideales")
    sheet.cell(row=initial_row + 3, column=2, value=np.mean(RUNS[0].valores_ideales))
    return workbook


def manage_refs(workbook):
    sheet = workbook.create_sheet("refs")
    sheet.cell(row=1, column=2, value="costo_por_agente")
    sheet.cell(row=1, column=3, value="std")
    sheet.cell(row=1, column=4, value="completion_time")
    sheet.cell(row=1, column=5, value="std")
    sheet.cell(row=1, column=6, value="valores_ideales")
    # sheet.cell(row=1, column=8, value="bad_pred")
    # sheet.cell(row=1, column=9, value="good_pred")
    # sheet.cell(row=1, column=10, value="total_pred")
    # sheet.cell(row=1, column=11, value="rate_pred")
    # sheet.cell(row=1, column=12, value="push_out_count")
    for i in range(len(events)):
        event = events[i]
        sheet.cell(row=i + 2, column=1, value=event)
        sheet.cell(row=i + 2, column=2, value="=" + event + "_agents!b43")
        sheet.cell(row=i + 2, column=3, value="=" + event + "_agents!d43")
        sheet.cell(row=i + 2, column=4, value="=" + event + "_agents!b86")
        sheet.cell(row=i + 2, column=5, value="=" + event + "_agents!d86")
        sheet.cell(row=i + 2, column=6, value="=" + event + "_agents!b133")

    initial_row = len(events) + 2
    sheet.cell(row=initial_row + 1, column=2, value="tiempo_ultimo_agente_goal")
    sheet.cell(row=initial_row + 1, column=3, value="MAX")
    sheet.cell(row=initial_row + 1, column=4, value="STDEV")
    for i in range(len(events)):
        event = events[i]
        sheet.cell(row=i + initial_row + 2, column=1, value=event)
        sheet.cell(row=i + initial_row + 2, column=2, value="=AVERAGE("+event+"_agents!b89:b128)")
        sheet.cell(row=i + initial_row + 2, column=3, value="=MAX("+event+"_agents!b89:b128)")
        sheet.cell(row=i + initial_row + 2, column=4, value="=STDEV("+event+"_agents!b89:b128)")
        # sheet.cell(row=i + initial_row + 2, column=2, value="=AVERAGE("+event+"_agents!b89:b128)")
        # sheet.cell(row=i + initial_row + 2, column=3, value="=AVERAGE("+event+"_agents!c89:c128)")
        # sheet.cell(row=i + initial_row + 2, column=4, value="=AVERAGE("+event+"_agents!e89:e128)")

    initial_row = (len(events) + 2) * 2
    sheet.cell(row=initial_row + 1, column=2, value="tiempo_en_acabar (success_agents)")
    sheet.cell(row=initial_row + 1, column=3, value="MAX")
    sheet.cell(row=initial_row + 1, column=4, value="STDEV")
    for i in range(len(events)):
        event = events[i]
        sheet.cell(row=i + initial_row + 2, column=1, value=event)
        sheet.cell(row=i + initial_row + 2, column=2, value="=AVERAGE("+event+"_agents!c89:c128)")
        sheet.cell(row=i + initial_row + 2, column=3, value="=MAX("+event+"_agents!c89:c128)")
        sheet.cell(row=i + initial_row + 2, column=4, value="=STDEV("+event+"_agents!c89:c128)")

    initial_row = (len(events) + 2) * 3
    sheet.cell(row=initial_row + 1, column=2, value="agentes_en_goal")
    sheet.cell(row=initial_row + 1, column=3, value="MAX")
    sheet.cell(row=initial_row + 1, column=4, value="STDEV")
    for i in range(len(events)):
        event = events[i]
        sheet.cell(row=i + initial_row + 2, column=1, value=event)
        sheet.cell(row=i + initial_row + 2, column=2, value="=AVERAGE("+event+"_agents!e89:e128)")
        sheet.cell(row=i + initial_row + 2, column=3, value="=MAX("+event+"_agents!e89:e128)")
        sheet.cell(row=i + initial_row + 2, column=4, value="=STDEV("+event+"_agents!e89:e128)")
    return workbook


if __name__ == '__main__':
    workbook = Workbook()
    for event in events:
        folder_name = event
        merge_file = open("./" + folder_name + "/" + merge_file_name, 'r')
        lines = strip_lines()
        RUNS = read_instances(lines)
        workbook = write_excel(RUNS, workbook)
    workbook = manage_refs(workbook)

    workbook.save(filename="resultsMARL.xlsx")
