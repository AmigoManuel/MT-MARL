from openpyxl import Workbook

folder_name = "100"
merge_file_name = "merge"


class RUN:
    def __init__(self, values):
        self.valores_ideales = values[0]
        self.costo_por_agente = values[1]
        self.completion_time_por_agente = values[2]
        self.tiempo_ultimo_agente_goal = values[3]
        self.tiempo_en_acabar = values[4]
        self.tiempo_promedio = values[5]
        self.agentes_en_goal = values[6]
        self.bad_good_total_rate_pred = values[7]


def strip_lines():
    lines = merge_file.readlines()
    for i in range(len(lines)):
        lines[i] = lines[i].strip()
    return lines


def read_instances(lines):
    RUNS = []
    temp = []
    for line in lines:
        name, values = line.split(' ')
        if name != "RUN":
            temp.append(eval(values))
        if name == "bad_good_total_rate_pred":
            RUNS.append(RUN(temp))
            temp.clear()
    return RUNS


def write_excel(RUNS, workbook):
    sheet = workbook.create_sheet(folder_name + " agents")

    sheet.cell(row=1, column=1, value="costo_por_agente")
    sheet.cell(row=2, column=1, value="RUN")
    for i in range(int(folder_name)):
        sheet.cell(row=2, column=i + 2, value="agent " + str(i + 1))
    for i in range(len(RUNS)):
        sheet.cell(row=i + 3, column=1, value=str(i))
        RUN = RUNS[i]
        for j in range(len(RUN.costo_por_agente)):
            sheet.cell(row=i + 3, column=j + 2, value=RUN.costo_por_agente[j])

    initial_row = len(RUNS) + 4
    sheet.cell(row=initial_row, column=1, value="completion_time_por_agente")
    sheet.cell(row=initial_row + 1, column=1, value="RUN")
    for i in range(int(folder_name)):
        sheet.cell(row=initial_row + 1, column=i + 2, value="agent " + str(i + 1))
    for i in range(len(RUNS)):
        sheet.cell(row=initial_row + i + 2, column=1, value=str(i))
        RUN = RUNS[i]
        for j in range(len(RUN.costo_por_agente)):
            sheet.cell(row=initial_row + i + 2, column=j + 2, value=RUN.costo_por_agente[j])

    initial_row = (initial_row * 2) - 1
    sheet.cell(row=initial_row, column=1, value="estadisticos")
    sheet.cell(row=initial_row + 1, column=1, value="RUN")
    sheet.cell(row=initial_row + 1, column=2, value="tiempo_ultimo_agente_goal")
    sheet.cell(row=initial_row + 1, column=3, value="tiempo_en_acabar")
    sheet.cell(row=initial_row + 1, column=4, value="tiempo_promedio")
    sheet.cell(row=initial_row + 1, column=5, value="agentes_en_goal")
    sheet.cell(row=initial_row + 1, column=6, value="bad_pred")
    sheet.cell(row=initial_row + 1, column=7, value="good_pred")
    sheet.cell(row=initial_row + 1, column=8, value="total_pred")
    sheet.cell(row=initial_row + 1, column=9, value="rate_pred")
    for i in range(len(RUNS)):
        RUN = RUNS[i]
        sheet.cell(row=initial_row + i + 2, column=1, value=str(i))
        sheet.cell(row=initial_row + i + 2, column=2, value=str(RUN.tiempo_ultimo_agente_goal))
        sheet.cell(row=initial_row + i + 2, column=3, value=str(RUN.tiempo_en_acabar))
        sheet.cell(row=initial_row + i + 2, column=4, value=str(RUN.tiempo_promedio))
        sheet.cell(row=initial_row + i + 2, column=5, value=str(RUN.agentes_en_goal))
        sheet.cell(row=initial_row + i + 2, column=6, value=str(RUN.bad_good_total_rate_pred[0]))
        sheet.cell(row=initial_row + i + 2, column=7, value=str(RUN.bad_good_total_rate_pred[1]))
        sheet.cell(row=initial_row + i + 2, column=8, value=str(RUN.bad_good_total_rate_pred[2]))
        sheet.cell(row=initial_row + i + 2, column=9, value=str(RUN.bad_good_total_rate_pred[3]))

    initial_row = initial_row + len(RUNS) + 3
    sheet.cell(row=initial_row, column=1, value="valores_ideales")
    for i in range(int(folder_name)):
        RUN = RUNS[0]
        sheet.cell(row=initial_row + 1, column=i + 1, value="agent " + str(i + 1))
        sheet.cell(row=initial_row + 2, column=i + 1, value=RUN.valores_ideales[i])

    return workbook


if __name__ == '__main__':
    events = ["50", "100", "150", "200", "250"]
    workbook = Workbook()
    for event in events:
        folder_name = event
        merge_file = open("./" + folder_name + "/" + merge_file_name, 'r')
        lines = strip_lines()
        RUNS = read_instances(lines)
        workbook = write_excel(RUNS, workbook)
    workbook.save(filename="results.xlsx")
